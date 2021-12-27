import json
import random
import threading
import time
from queue import Queue
from threading import Thread

import requests
from amazoncaptcha import AmazonCaptcha  # 用于识别反爬验证码
from bs4 import BeautifulSoup

from openpyxl import load_workbook

from amzon_spider.project.getInfo import getMovieInfo, getReviewInfo  # 用于解析网页

# +ASIN 为基本爬取网站
base_url = "https://www.amazon.com/dp/"

# 输入验证码获取页面的api 网址 get请求，需要参数params
error_url = "https://www.amazon.com/errors/validateCaptcha"

# # 参数格式
# params = {
#     'amzn': '',  # 验证码网页源码中
#     'amzn-r': '',  # 目标相对https://www.amazon.com的路径
#     'field-keywords': ''  # 验证码
# }

# 请求头user-agent
uaList = []


# 加载请求头user-agent
def loadUaList():
    wb = load_workbook("UAlist.xlsx")
    # 获取当前所有的sheet
    sheets = wb.worksheets
    # 获取第一张sheet
    sheet1 = sheets[0]
    # 通过Cell对象读取
    for i in range(2, 3500):
        uaList.append(sheet1.cell(i, 2).value)


# 随机获取一个代理ip
def get_proxy():
    return requests.get("http://127.0.0.1:5000/get/?type=http").json()


# 删除代理池中ip
def delete_proxy(proxy):
    requests.get("http://127.0.0.1:5000/delete/?proxy={}".format(proxy))


# 判断是否需要输入验证码
def is_code(html):
    soup = BeautifulSoup(html, "lxml")
    return soup.find(text="Enter the characters you see below") is not None


# 判断网页是否存在，即ASIN是否存在
def is_null(html):
    soup = BeautifulSoup(html, "lxml")
    return soup.find(name='img', attrs={
        'alt': "Sorry! We couldn't find that page. Try searching or go to Amazon's home page."}) is not None


# 判断页面是否缺少Product details块
def is_lack(html):
    soup = BeautifulSoup(html, "lxml")
    return soup.find(name='div', attrs={'id': "detailBulletsWrapper_feature_div"}) is None


# 破解验证码,获取参数
def get_params(html):
    soup = BeautifulSoup(html, "lxml")
    amzn = soup.find(name='input', attrs={'type': 'hidden', 'name': 'amzn'})['value']
    amzn_r = soup.find(name='input', attrs={'type': 'hidden', 'name': 'amzn-r'})['value']
    img_url = soup.find(name='div', attrs={'class': 'a-row a-text-center'}).find(name='img')['src']
    captcha = AmazonCaptcha.fromlink(img_url)
    code = captcha.solve()
    # 输入验证码获取页面的api 参数
    params = {
        'amzn': amzn,
        'amzn-r': amzn_r,
        'field-keywords': code
    }
    return params


# 爬虫代码
def get_html(ASIN, url):
    # 使用代理ip访问
    proxy = get_proxy().get("proxy")
    # 请求头
    header = {'user-agent': random.choice(uaList)}
    max_count = 3
    try:
        response = requests.get(url, headers=header, proxies={"http": "http://{}".format(proxy)})
        if is_null(response.text):
            return ASIN
        # 页面信息缺少Product details块
        while is_lack(response.text) and max_count > 0:
            proxy = get_proxy().get("proxy")
            header = {'user-agent': random.choice(uaList)}
            response = requests.get(url, headers=header, proxies={"http": "http://{}".format(proxy)})
            # 页面需要输入验证码
            while is_code(response.text):
                params = get_params(response.text)  # 获取验证码及参数
                time.sleep(random.randint(1, 3))  # 线程休眠,仿造输入验证码
                response = requests.get(error_url, headers=header, params=params, proxies={"http": "http://{}".format(proxy)})  # 发送验证码请求
            if is_null(response.text):
                return ASIN
            max_count -= 1
        return response.text
    except Exception:
        delete_proxy(proxy)
        return None


# 保存文件
def save(ASIN, html):
    # 访问失败
    if html is None:
        with open('../final/error.txt', 'a+', encoding='utf-8') as f0:
            f0.write(ASIN + '\n')
    else:
        # ASIN不存在
        if ASIN == html:
            with open('../final/unuseful.txt', 'a+', encoding='utf-8') as f0:
                f0.write(ASIN + '\n')
        # 爬取成功
        else:
            movie_info = getMovieInfo(ASIN, html)
            with open('../final/movie.json', 'a+', encoding='utf-8') as f1:
                f1.write(json.dumps(movie_info) + '\n')
            with open('../final/success.txt', 'a+', encoding='utf-8') as f0:
                f0.write(ASIN + '\n')

            # 评论数据
            reviews_info = getReviewInfo(ASIN, html)
            with open('../data/review.json', 'a+', encoding='utf-8') as f2:
                for review in reviews_info:
                    f2.write(json.dumps(review) + '\n')
            with open('../new/new_success.txt', 'a+', encoding='utf-8') as f0:
                f0.write(ASIN + '\n')


# 爬虫
def run(in_q, out_q):
    while in_q.empty() is not True:
        ASIN = in_q.get()
        url = base_url + ASIN
        html = get_html(ASIN, url)
        save(ASIN, html)
        time.sleep(random.random())
        out_q.put(str(threading.current_thread().getName()))
        in_q.task_done()


# 读取要爬取的ASIN
def getAsin(asin_queue):
    f = open('asin1.txt', encoding='utf-8')
    for line in f:
        asin_queue.put(line.strip())
    f.close()


if __name__ == '__main__':
    start_time = time.time()
    asin_queue = Queue()
    result_queue = Queue()

    # 读入所有asin，存入队列
    getAsin(asin_queue)

    # 加载请求头
    loadUaList()

    # 开始大小
    print('queue 开始大小 %d' % asin_queue.qsize())

    # 设置多线程
    for index in range(12):
        thread = Thread(target=run, args=(asin_queue, result_queue,))
        thread.daemon = True  # 随主线程退出而退出
        thread.start()

    asin_queue.join()  # 队列消费完 线程结束
    end_time = time.time()
    print('总耗时：%s' % (end_time - start_time))
    print('queue 结束大小 %d' % asin_queue.qsize())
    print('result_queue 结束大小 %d' % result_queue.qsize())
