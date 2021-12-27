from openpyxl import load_workbook

wb = load_workbook("../project/UAlist.xlsx")
sheets = wb.worksheets  # 获取当前所有的sheet
print(sheets)

# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
print(sheet1)


uaList = []
# 通过Cell对象读取
for i in range(2, 3500):
    uaList.append(sheet1.cell(i, 2).value)
print(uaList)

if __name__ == '__main__':
    print(uaList)
